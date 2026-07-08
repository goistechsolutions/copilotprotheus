"""
report_templates.py — Templates de estilização e layout para relatórios Excel e PDF.
Define paleta de cores (corporate navy/blue), fontes, alinhamentos e bordas.
"""
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.colors import HexColor
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.pdfgen import canvas

# ─── CORES CORPORATIVAS ───────────────────────────────────────────
COLOR_NAVY_DARK = "1E3A8A"  # #1e3a8a - Headers principais
COLOR_NAVY_LIGHT = "DBEAFE"  # #dbeafe - Totalizadores/Destaques
COLOR_GRAY_LIGHT = "F3F4F6"  # #f3f4f6 - Linhas alternadas
COLOR_TEXT_DARK = "1F2937"  # #1f2937 - Texto principal

# ─── EXCEL - ESTILOS (openpyxl) ───────────────────────────────────
font_title = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
font_body = Font(name="Calibri", size=11, color="000000")
font_total = Font(name="Calibri", size=11, bold=True, color="1E3A8A")

fill_title = PatternFill(start_color=COLOR_NAVY_DARK, end_color=COLOR_NAVY_DARK, fill_type="solid")
fill_header = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")  # Blue accent
fill_total = PatternFill(start_color=COLOR_NAVY_LIGHT, end_color=COLOR_NAVY_LIGHT, fill_type="solid")
fill_zebra = PatternFill(start_color=COLOR_GRAY_LIGHT, end_color=COLOR_GRAY_LIGHT, fill_type="solid")

align_center = Alignment(horizontal="center", vertical="center")
align_left = Alignment(horizontal="left", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")

thin_border_side = Side(border_style="thin", color="E5E7EB")
border_cell = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
border_total = Border(
    top=Side(border_style="thin", color="1E3A8A"),
    bottom=Side(border_style="double", color="1E3A8A")
)


def apply_auto_width(ws):
    """Ajusta automaticamente a largura das colunas do Excel com base no conteúdo."""
    for col in ws.columns:
        max_len = 0
        for cell in col:
            val = str(cell.value or '')
            if cell.number_format and ('R$' in cell.number_format or '%' in cell.number_format):
                val = f"R$ {val}"  # Adiciona margem de segurança para campos formatados
            max_len = max(max_len, len(val))
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)


# ─── PDF - ESTILOS (reportlab) ────────────────────────────────────

# Canvas numerado para gerar "Página X de Y" dinamicamente
class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_number(num_pages)
            canvas.Canvas.showPage(self)
        canvas.Canvas.save(self)

    def draw_page_number(self, page_count):
        self.saveState()
        self.setFont("Helvetica", 9)
        self.setFillColor(HexColor("#6B7280"))
        
        # Cabeçalho da página
        self.drawString(54, 750, "RODOL Ltda • Relatório Executivo Copilot Protheus")
        self.setStrokeColor(HexColor("#E5E7EB"))
        self.setLineWidth(0.5)
        self.line(54, 742, 558, 742)
        
        # Rodapé da página
        page_text = f"Página {self._pageNumber} de {page_count}"
        self.drawRightString(558, 40, page_text)
        self.drawString(54, 40, "Confidencial - Gerado eletronicamente")
        self.line(54, 52, 558, 52)
        self.restoreState()


def get_pdf_styles():
    """Retorna os estilos formatados para o ReportLab."""
    styles = getSampleStyleSheet()
    
    styles.add(ParagraphStyle(
        name='DocTitle',
        fontName='Helvetica-Bold',
        fontSize=20,
        leading=24,
        textColor=HexColor('#1E3A8A'),
        spaceAfter=15
    ))
    
    styles.add(ParagraphStyle(
        name='DocSubTitle',
        fontName='Helvetica',
        fontSize=10,
        leading=14,
        textColor=HexColor('#4B5563'),
        spaceAfter=20
    ))

    styles.add(ParagraphStyle(
        name='TableHeader',
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=11,
        textColor=HexColor('#FFFFFF'),
        alignment=1  # Center
    ))

    styles.add(ParagraphStyle(
        name='TableCell',
        fontName='Helvetica',
        fontSize=8,
        leading=10,
        textColor=HexColor('#1F2937')
    ))

    styles.add(ParagraphStyle(
        name='TableCellBold',
        fontName='Helvetica-Bold',
        fontSize=8,
        leading=10,
        textColor=HexColor('#1F2937')
    ))

    return styles


def get_table_style(has_zebra=True):
    """Estilo corporativo para tabelas ReportLab."""
    style_cmds = [
        ('BACKGROUND', (0, 0), (-1, 0), HexColor('#3B82F6')), # Blue accent para header
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#E5E7EB')),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
    ]
    if has_zebra:
        style_cmds.append(('ROWBACKGROUNDS', (0, 1), (-1, -1), [HexColor('#FFFFFF'), HexColor('#F9FAFB')]))
    
    return TableStyle(style_cmds)
