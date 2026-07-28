"""
report_service.py — Serviço de exportação de relatórios (Excel e PDF) para dados do Protheus.
Busca dados via Protheus REST API e gera os arquivos formatados.
"""
import os
import json
import uuid
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, List

from openpyxl import Workbook
from openpyxl.styles import Alignment

from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table
from reportlab.lib.pagesizes import letter

from app.services.protheus_service import execute_protheus_tool
from app.services import report_templates as tpl

# Diretório para armazenamento temporário dos relatórios gerados na nuvem
REPORTS_TMP_DIR = Path(os.getenv("REPORTS_TMP_DIR", "/tmp/copilot_reports"))
REPORTS_TMP_DIR.mkdir(parents=True, exist_ok=True)

# ─── CONFIGURAÇÃO DE RELATÓRIOS ───────────────────────────────────
REPORT_CONFIGS = {
    # Vendas / Faturamento
    "pedidos_abertos": {
        "title": "Pedidos de Venda Abertos",
        "endpoint": "PedidosRest",
        "headers": ["Documento", "Cliente", "Emissão", "Nota", "Bloqueado"],
        "fields": ["num", "cliente", "emissao", "nota", "bloquei"],
        "formats": {"emissao": "date"}
    },
    "itens_pedido": {
        "title": "Itens de Pedidos de Venda",
        "endpoint": "ItensPedidoRest",
        "headers": ["Pedido", "Item", "Produto", "Descrição", "Qtd. Vendida", "Qtd. Entregue", "Preço Unit."],
        "fields": ["num", "item", "produto", "descri", "qtdVen", "qtdEnt", "prcVen"],
        "formats": {"prcVen": "currency", "qtdVen": "number", "qtdEnt": "number"}
    },
    # Estoque
    "saldo_produtos": {
        "title": "Saldos de Estoque de Produtos",
        "endpoint": "SaldoRest",
        "headers": ["Cód. Produto", "Filial", "Qtd. Atual", "Qtd. Mín.", "Custo Méd."],
        "fields": ["B2_COD", "B2_FILIAL", "B2_QATU", "B2_QMIN", "B2_CM1"],
        "formats": {"B2_CM1": "currency", "B2_QATU": "number", "B2_QMIN": "number"}
    },
    # Fiscal
    "nfs_emitidas": {
        "title": "Notas Fiscais Emitidas",
        "endpoint": "NfsEmitidasRest",
        "headers": ["NF", "Série", "Cliente", "Loja", "Emissão", "Vl. Bruto", "Vl. Fiscal", "Chave NFe"],
        "fields": ["documento", "serie", "cliente", "loja", "emissao", "valorBruto", "valorFiscal", "chaveNfe"],
        "formats": {"emissao": "date", "valorBruto": "currency", "valorFiscal": "currency"}
    },
    "itens_nf": {
        "title": "Itens de Notas Fiscais (SFT)",
        "endpoint": "ItensNfRest",
        "headers": ["NF", "Série", "Item", "Produto", "Quantidade", "Preço Unit.", "Valor Cont."],
        "fields": ["nfiscal", "serie", "item", "produto", "quantidade", "precoUnit", "valorCont"],
        "formats": {"precoUnit": "currency", "valorCont": "currency", "quantidade": "number"}
    },
    "tes": {
        "title": "Tipos de Entrada e Saída (TES)",
        "endpoint": "TesRest",
        "headers": ["Código", "Texto", "Tipo", "CFOP", "ICMS", "IPI", "ISS"],
        "fields": ["codigo", "texto", "tipo", "cfop", "icms", "ipi", "iss"],
        "formats": {}
    },
    "livros_fiscais": {
        "title": "Livros Fiscais",
        "endpoint": "LivrosFiscaisRest",
        "headers": ["NF", "Série", "Cliente/Fornecedor", "Loja", "Emissão", "Base ICMS", "Vl. ICMS"],
        "fields": ["nfiscal", "serie", "clifor", "loja", "emissao", "baseIcm", "valorIcm"],
        "formats": {"emissao": "date", "baseIcm": "currency", "valorIcm": "currency"}
    },
    # Contábil
    "lancamentos": {
        "title": "Lançamentos Contábeis (CT2)",
        "endpoint": "LancamentosContabeisRest",
        "headers": ["Data", "Débito", "Crédito", "Valor", "Histórico", "CC Débito", "CC Crédito"],
        "fields": ["data", "debito", "credito", "valor", "historico", "ccDebito", "ccCredito"],
        "formats": {"data": "date", "valor": "currency"}
    },
    "balancete": {
        "title": "Balancete Contábil",
        "endpoint": "BalanceteRest",
        "headers": ["Conta", "Descrição", "Centro de Custo", "Custo Desc.", "Saldo Ant.", "Débitos", "Créditos", "Saldo Atual"],
        "fields": ["conta", "descricao", "centroCusto", "custoDesc", "saldoAnt", "debitos", "creditos", "saldoAtual"],
        "formats": {"saldoAnt": "currency", "debitos": "currency", "creditos": "currency", "saldoAtual": "currency"}
    },
    "plano_contas": {
        "title": "Plano de Contas",
        "endpoint": "PlanoContasRest",
        "headers": ["Cód. Custo", "Descrição", "Classe", "Bloqueado"],
        "fields": ["custo", "descricao", "classe", "bloq"],
        "formats": {}
    }
}


class ReportService:
    @staticmethod
    async def generate_report(report_type: str, filters: Dict[str, Any], file_format: str, tenant_id: str = "default") -> str:
        """
        Gera relatório com base no tipo, parâmetros e formato (xlsx ou pdf).
        Retorna o caminho físico do arquivo gerado.
        """
        if report_type not in REPORT_CONFIGS:
            raise ValueError(f"Tipo de relatório inválido: '{report_type}'")

        config = REPORT_CONFIGS[report_type]
        endpoint = config["endpoint"]

        # Normaliza cFilial para cFil para compatibilidade com as APIs ADVPL
        if "cFilial" in filters:
            filters["cFil"] = filters["cFilial"]
            # Mantém cFilial também para compatibilidade se necessário
            # filters.pop("cFilial")

        # Busca dados reais do Protheus
        res_raw = await execute_protheus_tool(endpoint, filters, tenant_id=tenant_id)
        res_json = json.loads(res_raw)

        if "error" in res_json:
            raise Exception(f"Erro Protheus: {res_json['error']}")

        items = []
        if isinstance(res_json, list):
            items = res_json
        elif isinstance(res_json, dict) and "items" in res_json:
            items = res_json["items"]

        # Gera UUID para arquivo único
        file_id = str(uuid.uuid4())
        ext = "xlsx" if file_format.lower() == "xlsx" else "pdf"
        file_path = REPORTS_TMP_DIR / f"{report_type}_{file_id}.{ext}"

        if ext == "xlsx":
            ReportService._build_xlsx(file_path, config, items)
        else:
            ReportService._build_pdf(file_path, config, items)

        return str(file_path)

    @staticmethod
    async def generate_from_markdown(markdown_content: str, file_format: str, title: str = "Relatorio Copilot") -> str:
        """
        Analisa o texto em markdown, extrai a primeira tabela encontrada
        e gera o arquivo Excel ou PDF correspondente.
        Retorna o caminho fisico do arquivo gerado.
        """
        import re
        
        # Procura a primeira tabela no formato markdown (| Col 1 | Col 2 |)
        lines = markdown_content.split("\n")
        table_rows = []
        in_table = False
        
        for line in lines:
            stripped = line.strip()
            if stripped.startswith("|") and stripped.endswith("|"):
                # Divide as celulas
                cells = [c.strip() for c in stripped.split("|")[1:-1]]
                
                # Se for linha separadora, pula
                if all(re.match(r"^:?-+:?$", cell) for cell in cells):
                    continue
                
                table_rows.append(cells)
                in_table = True
            else:
                if in_table and table_rows:
                    break
        
        if not table_rows or len(table_rows) < 2:
            raise ValueError("Nenhuma tabela valida encontrada no texto para exportacao.")
            
        headers = table_rows[0]
        data_rows = table_rows[1:]
        
        # Define fields como col_0, col_1, etc.
        fields = [f"col_{i}" for i in range(len(headers))]
        
        # Converte as linhas para objetos dict
        items = []
        for row in data_rows:
            # Completa a linha com vazio se faltar colunas
            row_completed = row + [""] * (len(headers) - len(row))
            item = {}
            for i, val in enumerate(row_completed[:len(headers)]):
                item[fields[i]] = val
            items.append(item)
            
        # Detecta formatos dinamicamente por amostragem das colunas
        formats = {}
        for i, field in enumerate(fields):
            column_values = [item[field] for item in items if item[field]]
            if not column_values:
                continue
                
            # Verifica se e coluna de data
            is_date = all(re.match(r"^\d{4}\d{2}\d{2}$", val) or re.match(r"^\d{2}/\d{2}/\d{4}$", val) for val in column_values)
            
            # Verifica se e coluna monetaria / R$
            is_currency = all("R$" in val or re.match(r"^\d+[\.,]\d{2}$", val.replace("R$","").strip()) for val in column_values)
            
            if is_date:
                formats[field] = "date"
            elif is_currency:
                formats[field] = "currency"
                # Limpa os caracteres monetarios
                for item in items:
                    raw_val = item[field]
                    if raw_val:
                        raw_val_clean = re.sub(r"[^\d\.,-]", "", raw_val).replace(".", "").replace(",", ".")
                        item[field] = raw_val_clean
        
        config = {
            "title": title,
            "headers": headers,
            "fields": fields,
            "formats": formats
        }
        
        # Gera o arquivo
        file_id = str(uuid.uuid4())
        ext = "xlsx" if file_format.lower() == "xlsx" else "pdf"
        file_path = REPORTS_TMP_DIR / f"export_{file_id}.{ext}"
        
        if ext == "xlsx":
            ReportService._build_xlsx(file_path, config, items)
        else:
            ReportService._build_pdf(file_path, config, items)
            
        return str(file_path)

    @staticmethod
    def _build_xlsx(file_path: Path, config: Dict[str, Any], items: List[Dict[str, Any]]):
        """Gera arquivo Excel (.xlsx) altamente formatado usando openpyxl."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório"
        ws.views.sheetView[0].showGridLines = True

        title = config["title"]
        headers = config["headers"]
        fields = config["fields"]
        formats = config.get("formats", {})

        # Título do Relatório
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=len(headers))
        title_cell = ws.cell(row=1, column=1, value=f"RODOL Ltda • {title}")
        title_cell.font = tpl.font_title
        title_cell.fill = tpl.fill_title
        title_cell.alignment = tpl.align_center

        # Espaçador
        ws.row_dimensions[3].height = 10

        # Cabeçalhos das Colunas
        ws.row_dimensions[4].height = 25
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=h)
            cell.font = tpl.font_header
            cell.fill = tpl.fill_header
            cell.alignment = tpl.align_center
            cell.border = tpl.border_cell

        # Linhas de Dados
        totalizers = {f: 0.0 for f in fields if formats.get(f) == "currency"}
        current_row = 5

        for item in items:
            ws.row_dimensions[current_row].height = 20
            is_zebra = (current_row % 2 == 0)

            for col_idx, f in enumerate(fields, 1):
                val = item.get(f, "")
                cell = ws.cell(row=current_row, column=col_idx)
                cell.border = tpl.border_cell

                if is_zebra:
                    cell.fill = tpl.fill_zebra

                fmt = formats.get(f)

                # Formatações numéricas e monetárias
                if fmt == "currency":
                    try:
                        num_val = float(val or 0)
                        cell.value = num_val
                        cell.number_format = '"R$"#,##0.00'
                        cell.alignment = tpl.align_right
                        totalizers[f] += num_val
                    except (ValueError, TypeError):
                        cell.value = val
                        cell.alignment = tpl.align_left
                elif fmt == "number":
                    try:
                        cell.value = float(val or 0) if "." in str(val) else int(val or 0)
                        cell.alignment = tpl.align_right
                    except (ValueError, TypeError):
                        cell.value = val
                        cell.alignment = tpl.align_left
                elif fmt == "date":
                    # Formata YYYYMMDD para DD/MM/YYYY se aplicável
                    if len(str(val)) == 8:
                        cell.value = f"{val[6:8]}/{val[4:6]}/{val[0:4]}"
                    else:
                        cell.value = val
                    cell.alignment = tpl.align_center
                else:
                    cell.value = val
                    cell.alignment = tpl.align_left
                    cell.font = tpl.font_body

            current_row += 1

        # Adiciona totalizadores se houver colunas financeiras
        if any(totalizers.values()):
            ws.row_dimensions[current_row].height = 22
            ws.cell(row=current_row, column=1, value="TOTAL GERAL").font = tpl.font_total
            
            for col_idx, f in enumerate(fields, 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.fill = tpl.fill_total
                cell.border = tpl.border_total
                
                if formats.get(f) == "currency":
                    cell.value = totalizers[f]
                    cell.number_format = '"R$"#,##0.00'
                    cell.alignment = tpl.align_right
                    cell.font = tpl.font_total

        tpl.apply_auto_width(ws)
        wb.save(str(file_path))

    @staticmethod
    def _build_pdf(file_path: Path, config: Dict[str, Any], items: List[Dict[str, Any]]):
        """Gera arquivo PDF (.pdf) executivo altamente formatado usando reportlab."""
        styles = tpl.get_pdf_styles()

        # Configuração do Documento
        doc = SimpleDocTemplate(
            str(file_path),
            pagesize=letter,
            rightMargin=54,
            leftMargin=54,
            topMargin=72,
            bottomMargin=72
        )

        elements = []

        # Título do Relatório
        elements.append(Paragraph(f"RODOL Ltda • {config['title']}", styles['DocTitle']))
        
        # Subtítulo com metadados
        gen_time = datetime.now().strftime("%d/%m/%Y às %H:%M:%S")
        elements.append(Paragraph(f"Relatório gerado em {gen_time} | Total de registros: {len(items)}", styles['DocSubTitle']))
        elements.append(Spacer(1, 10))

        headers = config["headers"]
        fields = config["fields"]
        formats = config.get("formats", {})

        # Dados da Tabela
        table_data = []
        
        # Header da Tabela
        header_row = [Paragraph(h, styles['TableHeader']) for h in headers]
        table_data.append(header_row)

        totalizers = {f: 0.0 for f in fields if formats.get(f) == "currency"}

        # Linhas da Tabela
        for item in items:
            row = []
            for f in fields:
                val = item.get(f, "")
                fmt = formats.get(f)
                
                if fmt == "currency":
                    try:
                        num_val = float(val or 0)
                        formatted_val = f"R$ {num_val:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                        row.append(Paragraph(formatted_val, styles['TableCell']))
                        totalizers[f] += num_val
                    except (ValueError, TypeError):
                        row.append(Paragraph(str(val), styles['TableCell']))
                elif fmt == "date" and len(str(val)) == 8:
                    formatted_val = f"{val[6:8]}/{val[4:6]}/{val[0:4]}"
                    row.append(Paragraph(formatted_val, styles['TableCell']))
                else:
                    row.append(Paragraph(str(val), styles['TableCell']))
            table_data.append(row)

        # Adiciona linha de total se houver valores monetários
        if any(totalizers.values()):
            total_row = []
            for idx, f in enumerate(fields):
                if idx == 0:
                    total_row.append(Paragraph("<b>TOTAL GERAL</b>", styles['TableCellBold']))
                elif formats.get(f) == "currency":
                    num_val = totalizers[f]
                    formatted_val = f"R$ {num_val:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                    total_row.append(Paragraph(f"<b>{formatted_val}</b>", styles['TableCellBold']))
                else:
                    total_row.append(Paragraph("", styles['TableCell']))
            table_data.append(total_row)

        # Criação da Tabela com dimensões flexíveis
        col_width = (doc.width) / len(headers)
        t = Table(table_data, colWidths=[col_width] * len(headers))
        t.setStyle(tpl.get_table_style())
        elements.append(t)

        # Constrói o PDF usando o NumberedCanvas customizado para rodapés dinâmicos
        doc.build(elements, canvasmaker=tpl.NumberedCanvas)
